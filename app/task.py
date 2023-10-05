import pandas as pd
import datetime 
import openpyxl

# Carregar o arquivo Excel
input_workbook = openpyxl.load_workbook(r"C:\Users\RafaelSouza\VERUM PARTNERS\VERUM PARTNERS - VAL2018021\00.TI\Proj - New Steel\BI\01. Dashboards Ativos\Squad Manuseio\2023-08-25 - painel fluxo engenharia - fornecimento - construção transportadores de correia.xlsx")

# Escolher a planilha desejada
input_sheet = input_workbook['Fluxo civil']

# Criar um novo arquivo Excel para os comentários
output_workbook = openpyxl.Workbook()
output_sheet = output_workbook.active
output_sheet.title = 'Comentarios'

# Iterar sobre as células da planilha
for row in input_sheet.iter_rows():
    for cell in row:
        # Verificar se a célula possui um comentário
        if cell.comment:
            author = cell.comment.author  # Autor do comentário
            text = cell.comment.text  # Texto do comentário
            cell_index = cell.coordinate  # Índice da célula

            # Obter a célula equivalente na nova planilha e inserir o comentário
            output_cell = output_sheet[cell_index]
            output_cell.value = f'Autor: {author}\nComentário: {text}'

            #print(f'Célula {cell_index} {cell.coordinate} - Autor: {author}, Comentário: {text}')


# Salvar o novo arquivo Excel com os comentários
output_workbook.save(r"C:\Users\RafaelSouza\VERUM PARTNERS\VERUM PARTNERS - VAL2018021\00.TI\Proj - New Steel\BI\01. Dashboards Ativos\Squad Manuseio\comentarios.xlsx")

# Fechar os arquivos Excel
input_workbook.close()
output_workbook.close()
